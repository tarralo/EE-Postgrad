'''Trabalho 1 - Implementação de Métodos Iterativos para Análise de Sistemas de Potência
   Autor: Luiz A. Tarralo
   Data: 10/04/2023'''

import numpy as np
import pandas as pd
import math 
import cmath
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import tkinter as tk
from pandastable import Table

# função main
def main():
    
    data_bus, data_branch = read_data() # chama função para leitura dos dados
    num_bus = len(data_bus) # número de barras
    Ym = calc_Y(data_branch, num_bus) # chama função para cálculo da matriz de admitância

    window = tk.Tk() # cria a janela
    window.title('Análise de Sistemas de Potência por Métodos Iterativos')
    window.geometry('800x800')
    window.resizable(True, True)

    autor = tk.Label(window, text='Criado por Luiz A. Tarralo', font=('Arial', 10, 'bold'))
    autor.pack(side='top', anchor='nw')

    V_gs, S_gs, it_gs, erro_gs = gs_method(data_bus, Ym, num_bus, 1e-6, 100) # chama função para método de Gauss-Seidel
    data_busGS = update_data(data_bus, V_gs, S_gs) # atualiza dataframe de barras com os resultados do método de Gauss-Seidel
    tk.Label(window, text=f'MÉTODO GAUSS-SEIDEL \n', font=('Arial', 12, 'bold')).pack() # mostra o método utilizado
    tk.Label(window, text=f'{it_gs} iterações\nErro = {erro_gs}').pack() # mostra o número de iterações e o erro
    gs = tk.Frame(window) # cria um frame para a tabela do método de Gauss-Seidel
    gs.pack(fill='both', expand=False, side='top') # mostra o frame
    gs_table = Table(gs, dataframe=data_busGS, showtoolbar=False, showstatusbar=True) # cria a tabela
    gs_table.show() # mostra a tabela

    V_nr, S_nr, it_nr, erro_nr = nr_method(data_bus, Ym, num_bus, 1e-3, 100) # chama função para método de Newton-Raphson
    print(V_nr, S_nr)
    data_busNR = update_data(data_bus, V_nr, S_nr) # atualiza dataframe de barras com os resultados do método de Newton-Raphson
    tk.Label(window, text=f'MÉTODO NEWTON-RAPHSON \n', font=('Arial', 12, 'bold')).pack() # mostra o método utilizado
    tk.Label(window, text=f'{it_nr} iterações\nErro = {erro_nr}').pack() # mostra o número de iterações e o erro
    nr = tk.Frame(window) # cria um frame para a tabela do método de Newton-Raphson
    nr.pack(fill='both', expand=False, side='bottom') # mostra o frame
    nr_table = Table(nr, dataframe=data_busNR, showtoolbar=True, showstatusbar=True) # cria a tabela
    nr_table.show() # mostra a tabela

    save_excel(data_busGS, data_busNR, it_gs, erro_gs, it_nr, erro_nr) # chama função para salvar os resultados em um arquivo excel

    window.mainloop() # mostra a interface gráfica
    
# função para leitura dos dados
def read_data():

    # leitura dos dados do arquivo
    data_bus = pd.read_excel('data.xlsx', sheet_name='bus', index_col='number')
    data_branch = pd.read_excel('data.xlsx', sheet_name='branch')

    # preenche NaN com 0.0
    data_bus = data_bus.fillna(0.0)
    data_branch = data_branch.fillna(0.0)

    return data_bus, data_branch

# função para cálculo da matriz de admitância
def calc_Y(data_branch, num_bus):

    Ym = np.zeros((num_bus, num_bus), dtype=complex) # inicializa matriz de admitância

    for i in data_branch.index: # itera sobre as linhas do dataframe de ramo

        from_b, to_b = int(data_branch.iloc[i]['from']), int(data_branch.iloc[i]['to']) # barras de origem e destino
        r, x, b = data_branch.iloc[i]['r'], data_branch.iloc[i]['xl'], data_branch.iloc[i]['b'] # resistência, reatância e susceptância
        Ym[from_b][from_b] += (1/complex(r,x)) + complex(0,b/2) # diagonal principal
        Ym[to_b][to_b] += (1/complex(r,x)) + complex(0,b/2) # diagonal principal
        Ym[from_b][to_b] += -1/complex(r,x) # diagonal secundária
        Ym[to_b][from_b] += -1/complex(r,x) # diagonal secundária

    return Ym

# função para valores iniciais
def init_values(data_bus, num_bus):

    V_init = np.zeros(num_bus, dtype=complex) # tensão inicial
    S_init = np.zeros(num_bus, dtype=complex) # potência inicial

    ref_index = data_bus.loc[data_bus['type'] == 'SLACK'].index[0] # índice da barra de referência
    ref_angle = math.radians(data_bus.iloc[ref_index]['angle']) # ângulo da barra de referência em radianos

    for i in data_bus.index:

        if data_bus.iloc[i]['type'] == 'SLACK' or data_bus.iloc[i]['type'] == 'PV': 
            V_init[i] = cmath.rect(data_bus.iloc[i]['voltage'], ref_angle) # converte para retangular

        else:
            V_init[i] = cmath.rect(1.0,ref_angle) # converte para retangular com tensão 1.0 p.u.

        S_init[i] = complex(data_bus.iloc[i]['pg'] - data_bus.iloc[i]['pl'], data_bus.iloc[i]['qg'] - data_bus.iloc[i]['ql']) # potência complexa inicial

    return V_init, S_init, ref_index

# método de Gauss-Seidel
def gs_method(data_bus, Ym, num_bus, tol, max_iter):

    tensao, potencia, ref_index = init_values(data_bus, num_bus) # chama função para valores iniciais
    erro = 1 # erro inicial
    it = 0 # inicia número de iterações

    while erro > tol: # primeiro critério de parada, erro menor que a tolerância

        V_init = np.copy(tensao) # armazena tensão anterior

        for i in data_bus.index: # itera sobre as barras

            if data_bus.iloc[i]['type'] == 'PV': # calcula Qi para barras PV

                corrente = complex(0,0) # inicializa corrente

                for j in data_bus.index:

                    corrente += tensao[j] * Ym[i][j] # soma correntes pela lei de Kirchhoff

                qi = - (tensao[i].conjugate() * corrente).imag # potência reativa injetada
                potencia [i] = complex(potencia[i].real, qi) # atualiza potência complexa inicial

            if data_bus.iloc[i]['type'] != 'SLACK': # calcula tensão para barras PQ

                corrente = complex(0,0) # inicializa corrente

                for j in data_bus.index:

                    if i != j:

                        corrente += tensao[j] * Ym[i][j] # soma correntes pela lei de Kirchhoff

                tensao[i] = (1/Ym[i][i]) * (potencia[i].conjugate() / tensao[i].conjugate() - corrente) # atualiza tensão inicial

                if data_bus.iloc[i]['type'] == 'PV':

                    tensao[i] = cmath.rect(data_bus.iloc[i]['voltage'], cmath.phase(tensao[i])) # mantém o módulo e atualiza o ângulo

        erro = np.amax(abs(abs(tensao) - abs(V_init))) # atualiza erro
        it += 1 # atualiza número de iterações

        if it > max_iter: # segundo critério de parada, número máximo de iterações
            print(f'MÉTODO GAUSS-SEIDEL \nNúmero máximo de iterações atingido: {max_iter}\n Erro: {erro:.6f}\n NÃO CONVERGIU!')
            break

    # cálculo da potência complexa injetada na barra de referência
    corrente = complex(0,0) # reinicializa corrente
    for i in data_bus.index:
        corrente += tensao[i] * Ym[ref_index][i] 
    potencia[ref_index] = tensao[ref_index] * corrente.conjugate() # potência complexa injetada na barra de referência

    return tensao, potencia, it, erro

# método de Newton-Raphson
def nr_method(data_bus, Ym, num_bus, tol, max_iter):

    tensao, potencia, ref_index = init_values(data_bus, num_bus) # chama função para valores iniciais
    
    npv = 0 # número de barras PV
    npq = 0 # número de barras PQ
    for i in data_bus.index:
        if data_bus.iloc[i]['type'] == 'PV':
            npv += 1
        elif data_bus.iloc[i]['type'] == 'PQ':
            npq += 1
    
    d_pq = np.zeros(2 * npq + npv) # vetor de potências
    potencia_esp = np.copy(potencia) # vetor de potências esperadas

    erro = 1 # erro inicial
    it = 0 # inicia número de iterações

    while erro > tol: 
        nump = 0 
        numq = 0
        for i in data_bus.index:
            if i != ref_index: # não considera a barra de referência
                corrente = complex(0,0) # inicializa corrente
                for j in data_bus.index:
                    corrente += Ym[i][j] * tensao[j] # soma correntes pela lei de Kirchhoff
    
                potencia[i] = tensao[i] * corrente.conjugate() # potência complexa injetada na barra i
                d_pq [nump] = potencia_esp[i].real - potencia[i].real # delta da potência ativa
                nump += 1

                if data_bus.iloc[i]['type'] == 'PQ': 
                    d_pq [numq + npq + npv] = potencia_esp[i].imag - potencia[i].imag # delta da potência reativa
                    numq += 1

        erro = np.amax(abs(d_pq)) # atualiza erro
        if erro < tol:
            break

        # calcula a matriz Jacobiana
        Jm = jacobiana(tensao, potencia, Ym, data_bus, npv, npq, ref_index)

        # resolve o sistema de equações
        delta_VT = np.linalg.solve(Jm, d_pq)

        # atualiza tensões
        num_T, num_V = 0, 0
        for i in data_bus.index:
            
            if data_bus.iloc[i]['type'] == 'PQ': # atualiza módulo e ângulo para barras PQ
                tensao[i] = cmath.rect(abs(tensao[i]) + delta_VT[num_T + npq + npv], cmath.phase(tensao[i]) + delta_VT[num_V]) # atualiza módulo e ângulo
                num_T += 1
                num_V += 1

            elif data_bus.iloc[i]['type'] == 'PV': # atualiza ângulo para barras PV
                tensao[i] = cmath.rect(abs(tensao[i]), cmath.phase(tensao[i]) + delta_VT[num_V]) # atualiza ângulo
                num_V += 1

        it += 1 # atualiza número de iterações

        if it > max_iter: # segundo critério de parada, número máximo de iterações
            print(f'MÉTODO NEWTON-RAPHSON \nNúmero máximo de iterações atingido: {max_iter}\nErro: {erro:.6f}\nNÃO CONVERGIU!')
            break
           
    # atualiza potência complexa injetada na barra de referência
    corrente = complex(0,0) # reinicializa corrente
    for i in data_bus.index:
        corrente += Ym[ref_index][i] * tensao[i]
    potencia[ref_index] = tensao[ref_index] * corrente.conjugate() # potência complexa injetada na barra de referência

    return tensao, potencia, it, erro

# função para calcular a matriz Jacobiana
def jacobiana(tensao, potencia, Ym, data_bus, npv, npq, ref_index):
    j1 = np.zeros((npq + npv, npq + npv)) # inicializa matriz j1
    j2 = np.zeros((npq + npv, npq)) # inicializa matriz j2
    j3 = np.zeros((npq, npq + npv)) # inicializa matriz j3
    j4 = np.zeros((npq, npq)) # inicializa matriz j4
    bus_index = np.delete(data_bus.index.to_numpy(), ref_index) # vetor com índices das barras

    # calcula a matriz j1
    j1_i = {'l': 0, 'c': 0} # inicializa índices da matriz j1
    for i in bus_index: # num_pq + num_pv linhas
        j1_i['c'] = 0
        for j in bus_index: # num_pq + num_pv colunas
            if i == j:
                j1[j1_i['l'], j1_i['c']] = - potencia[i].imag - pow(abs(tensao[i]),2) * Ym[i][i].imag 
            else:
                j1[j1_i['l'], j1_i['c']] = abs(tensao[i]) * abs(tensao[j]) * (Ym[i][j].real * math.sin(cmath.phase(Ym[i][j])) - Ym[i][j].imag * math.cos(cmath.phase(Ym[i][j])))
            j1_i['c'] += 1 # atualiza coluna
        j1_i['l'] += 1 # atualiza linha

    # calcula a matriz j2
    j2_i = {'l':0, 'c':0} # inicializa índices da matriz j2
    for i in bus_index: # num_pq + num_pv linhas
        j2_i['c'] = 0
        for j in bus_index: # num_pq colunas
            if data_bus.iloc[j]['type'] == 'PQ':
                if i == j: 
                    j2[j2_i['l'],j2_i['c']] = (potencia[i].real + pow(abs(tensao[i]),2) * Ym[i][i].real) / (abs(tensao[i].real))
                else:
                    j2[j2_i['l'],j2_i['c']] = abs(tensao[i]) * (Ym[i][j].real * math.cos(cmath.phase(Ym[i][j])) + Ym[i][j].imag * math.sin(cmath.phase(Ym[i][j])))
                j2_i['c'] += 1 # atualiza coluna
        j2_i['l'] += 1 # atualiza linha
    
    # calcula a matriz j3
    j3_i = {'l':0, 'c':0} # inicializa índices da matriz j3
    for i in bus_index: 
        if data_bus.iloc[i]['type'] == 'PQ':
            j3_i['c'] = 0
            for j in bus_index: # num_pq + num_pv colunas
                if i == j:
                    j3[j3_i['l'],j3_i['c']] = potencia[i].real - pow(abs(tensao[i]),2) * Ym[i][i].real
                else:
                    j3[j3_i['l'],j3_i['c']] = - abs(tensao[i]) * abs(tensao[j]) * (Ym[i][j].real * math.cos(cmath.phase(Ym[i][j])) + Ym[i][j].imag * math.sin(cmath.phase(Ym[i][j])))
                j3_i['c'] += 1 # atualiza coluna
            j3_i['l'] += 1 # atualiza linha

    # calcula a matriz j4
    j4_i = {'l':0, 'c':0} # inicializa índices da matriz j4
    for i in bus_index:
        if data_bus.iloc[i]['type'] == 'PQ': # num_pq linhas
            j4_i['c'] = 0 # reinicia coluna
            for j in bus_index:
                if data_bus.iloc[j]['type'] == 'PQ': # num_pq colunas
                    if i == j:
                        j4[j4_i['l'],j4_i['c']] = (potencia[i].imag - pow(abs(tensao[i]),2) * Ym[i][i].imag) / (abs(tensao[i])) 
                    else:
                        j4[j4_i['l'],j4_i['c']] = abs(tensao[i]) * (Ym[i][j].real * math.sin(cmath.phase(Ym[i][j])) - Ym[i][j].imag * math.cos(cmath.phase(Ym[i][j])))
                    j4_i['c'] += 1 # atualiza coluna
            j4_i['l'] += 1 # atualiza linha
        
    Jm = np.concatenate((np.concatenate((j1, j2), axis=1), np.concatenate((j3, j4), axis=1)), axis=0)

    return Jm

# função para atualizar os dados da barra
def update_data(data_bus, tensao, potencia):

    # atualiza tensões e potências obtidas
    data_bus['voltage'] = abs(tensao)
    data_bus['angle'] = np.angle(tensao, deg=True) 
    for i in data_bus.index:
        if data_bus.iloc[i]['type'] == 'PV':
            data_bus.at[i, 'qg'] = potencia[i].imag - data_bus.iloc[i]['ql'] # atualiza potência reativa gerada

        elif data_bus.iloc[i]['type'] == 'SLACK':
            data_bus.at[i, 'pg'] = potencia[i].real - data_bus.iloc[i]['pl'] # atualiza potência ativa gerada
            data_bus.at[i, 'qg'] = potencia[i].imag - data_bus.iloc[i]['ql'] # atualiza potência reativa gerada

    return data_bus

# função para atualizar os dados e passar a um arquivo .xlsx
def save_excel(data_busGS, data_busNR, it_gs, erro_gs, it_nr, erro_nr):

    with pd.ExcelWriter('results.xlsx') as writer:

        # adiciona os dados do Gauss-Seidel 
        data_busGS.to_excel(writer, sheet_name='Gauss-Seidel', float_format = "%.12f", index=False) # adiciona os dados do Gauss-Seidel
        worksheet_gs = writer.sheets['Gauss-Seidel'] # cria uma worksheet
        linha_gs = len(data_busGS) + 2 # linha para escrever o número de iterações e o erro
        #worksheet_gs.cell(row=linha_gs, column=1).value = f"Convergiu em {it_gs} iterações com um erro de {erro_gs:.6e}."

        # adiciona os dados do Newton-Raphson
        data_busNR.to_excel(writer, sheet_name='Newton-Raphson', float_format = "%.12f", index=False) # adiciona os dados do Newton-Raphson
        worksheet_nr = writer.sheets['Newton-Raphson'] # cria uma worksheet
        linha_nr = len(data_busNR) + 2 # linha para escrever o número de iterações e o erro
        #worksheet_nr.cell(row=linha_nr, column=1).value = f"Convergiu em {it_nr} iterações com um erro de {erro_nr:.6e}."

        #for sheet in writer.sheets.values():
        #    sheet.column_dimensions[get_column_letter(1)].width = 40 # ajusta a largura da coluna 1

# chama função main
if __name__ == "__main__":
    main()







